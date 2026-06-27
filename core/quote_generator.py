"""
見積書生成モジュール
Excel（openpyxl）とPDF（reportlab）で会社フォーマットの見積書を出力する
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Optional


OUTPUT_DIR = Path(__file__).parent.parent / "output"


class QuoteGenerator:
    """
    見積書生成クラス
    Excel・PDFフォーマットで見積書を出力する
    """

    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir) if output_dir else OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ─────────────────────────────────────────
    # テンプレート流し込み方式（メイン）
    # ─────────────────────────────────────────
    def generate_from_template(
        self,
        template_id: str,
        estimation: dict,
        project_data: dict,
        client_name: str = "",
        site_address: str = "",
        sales_rep: str = "",
        company_name: str = "",
        discount: int = 0,
    ) -> str:
        """
        登録済みテンプレートにAI積算結果を流し込んでExcelを生成する

        Args:
            template_id: テンプレートID（例: "standard"）
            estimation:  EstimationEngineの出力
            ...
        Returns:
            str: 出力ファイルパス
        """
        from core.template_manager import copy_template_to_output
        from core.template_filler import fill_template

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        label = client_name or "案件"
        filename = f"見積書_{label}_{ts}.xlsx"

        # テンプレートを出力ディレクトリにコピー
        output_path = copy_template_to_output(template_id, self.output_dir, filename)
        if output_path is None:
            raise FileNotFoundError(f"テンプレート '{template_id}' が見つかりません")

        # セルに値を流し込む
        from core.template_manager import get_template_path
        template_path = get_template_path(template_id)

        fill_template(
            template_id=template_id,
            template_path=template_path,
            output_path=output_path,
            estimation=estimation,
            project_data=project_data,
            client_name=client_name,
            site_address=site_address,
            sales_rep=sales_rep,
            company_name=company_name,
            discount=discount,
        )
        return str(output_path)

    # ─────────────────────────────────────────
    # Excel出力
    # ─────────────────────────────────────────
    def generate_excel(
        self,
        estimation: dict,
        project_data: dict,
        client_name: str = "",
        site_address: str = "",
        sales_rep: str = "",
    ) -> str:
        """
        Excelフォーマットの見積書を生成する

        Returns:
            str: 出力ファイルパス
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment, Border, Font, PatternFill, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxlをインストールしてください: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "見積書"

        # ── スタイル定義 ──
        NAVY = "1A237E"
        LIGHT_BLUE = "E3F2FD"
        LIGHT_GREEN = "E8F5E9"
        ORANGE = "FFF3E0"
        GRAY = "F5F5F5"

        def cell_style(ws, row, col, value="", bold=False, bg=None, align="left",
                       font_size=11, border=True, wrap=False, number_format=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name="游ゴシック", size=font_size, bold=bold,
                          color="FFFFFF" if bg == NAVY else "000000")
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            if border:
                side = Side(style="thin", color="CCCCCC")
                c.border = Border(left=side, right=side, top=side, bottom=side)
            if number_format:
                c.number_format = number_format
            return c

        # ── 列幅設定 ──
        col_widths = {1: 5, 2: 20, 3: 30, 4: 10, 5: 8, 6: 14, 7: 14, 8: 10, 9: 30}
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # ── 行高さ設定 ──
        for row in range(1, 60):
            ws.row_dimensions[row].height = 20

        row = 1

        # ── タイトルブロック ──
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1, value="御　見　積　書")
        c.font = Font(name="游ゴシック", size=20, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 40
        row += 1

        # ── 案件情報 ──
        today = datetime.now().strftime("%Y年%m月%d日")
        info = [
            ("発行日",    today),
            ("お客様名",  f"{client_name} 御中" if client_name else ""),
            ("現場住所",  site_address or ""),
            ("担当者",    sales_rep or ""),
        ]
        for label, value in info:
            ws.merge_cells(f"A{row}:B{row}")
            cell_style(ws, row, 1, label, bold=True, bg=LIGHT_BLUE, align="center")
            ws.merge_cells(f"C{row}:I{row}")
            cell_style(ws, row, 3, value)
            row += 1

        row += 1

        # ── 合計金額ブロック ──
        ws.merge_cells(f"A{row}:C{row}")
        cell_style(ws, row, 1, "御見積合計金額（税込）", bold=True, bg=NAVY, align="center", font_size=12)
        ws.merge_cells(f"D{row}:I{row}")
        total_val = estimation.get("total", 0)
        c = ws.cell(row=row, column=4, value=total_val)
        c.font = Font(name="游ゴシック", size=16, bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        c.number_format = '¥#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 32
        row += 2

        # ── 明細ヘッダー ──
        headers = ["No", "工種", "品目", "数量", "単位", "単価（円）", "金額（円）", "状態", "備考"]
        bg_colors = [NAVY] * len(headers)
        for col, (header, bg) in enumerate(zip(headers, bg_colors), 1):
            cell_style(ws, row, col, header, bold=True, bg=bg, align="center", font_size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        # ── 明細行 ──
        items = estimation.get("estimation_items", [])
        for i, item in enumerate(items, 1):
            estimated = item.get("estimated", False)
            needs_conf = item.get("needs_confirmation", False)
            row_bg = ORANGE if needs_conf else (LIGHT_BLUE if estimated else None)

            cell_style(ws, row, 1, i, align="center", bg=row_bg)
            cell_style(ws, row, 2, item.get("category", ""), bg=row_bg)
            cell_style(ws, row, 3, item.get("item_name", ""), bg=row_bg)
            cell_style(ws, row, 4, item.get("quantity", 0), align="right", bg=row_bg)
            cell_style(ws, row, 5, item.get("unit", ""), align="center", bg=row_bg)

            c_price = ws.cell(row=row, column=6, value=item.get("unit_price", 0))
            c_price.number_format = '#,##0'
            c_price.alignment = Alignment(horizontal="right")
            if row_bg:
                c_price.fill = PatternFill("solid", fgColor=row_bg)

            c_amount = ws.cell(row=row, column=7, value=item.get("amount", 0))
            c_amount.number_format = '#,##0'
            c_amount.alignment = Alignment(horizontal="right")
            if row_bg:
                c_amount.fill = PatternFill("solid", fgColor=row_bg)

            status = "⚠️要確認" if needs_conf else ("📊推定" if estimated else "✅確定")
            cell_style(ws, row, 8, status, align="center", bg=row_bg)
            cell_style(ws, row, 9, item.get("basis", "") or item.get("notes", ""),
                       bg=row_bg, wrap=True, font_size=9)
            row += 1

        # ── 小計・消費税・合計 ──
        row += 1
        summary = [
            ("小計（税抜）",  estimation.get("subtotal", 0),   None),
            ("消費税（10%）", estimation.get("tax_amount", 0),  None),
            ("合計（税込）",  estimation.get("total", 0),       LIGHT_GREEN),
        ]
        for label, value, bg in summary:
            ws.merge_cells(f"A{row}:F{row}")
            cell_style(ws, row, 1, label, bold=True, align="right",
                       bg=bg or GRAY, border=True)
            c = ws.cell(row=row, column=7, value=value)
            c.font = Font(name="游ゴシック", bold=True,
                          size=13 if label == "合計（税込）" else 11)
            c.number_format = '¥#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center")
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            ws.merge_cells(f"H{row}:I{row}")
            row += 1

        # ── 注記 ──
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1,
                    value="※ 📊推定：現場写真・説明から算出した推定値です。現地確認後に変更になる場合があります。")
        c.font = Font(name="游ゴシック", size=9, color="E65100")
        ws.row_dimensions[row].height = 16
        row += 1

        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1,
                    value="※ ⚠️要確認：現場での確認が必要な項目です。")
        c.font = Font(name="游ゴシック", size=9, color="C62828")

        # ── ファイル保存 ──
        filename = self.output_dir / f"見積書_{client_name or '案件'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(str(filename))
        return str(filename)

    # ─────────────────────────────────────────
    # PDF出力
    # ─────────────────────────────────────────
    def generate_pdf(
        self,
        estimation: dict,
        project_data: dict,
        client_name: str = "",
        site_address: str = "",
        sales_rep: str = "",
    ) -> str:
        """
        PDFフォーマットの見積書を生成する

        Returns:
            str: 出力ファイルパス
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            from reportlab.platypus import (
                Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
            )
        except ImportError:
            raise ImportError("reportlabをインストールしてください: pip install reportlab")

        # 日本語フォント設定
        try:
            pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
            JP_FONT = "HeiseiKakuGo-W5"
        except Exception:
            JP_FONT = "Helvetica"  # フォールバック

        filename = self.output_dir / f"見積書_{client_name or '案件'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        doc = SimpleDocTemplate(
            str(filename),
            pagesize=A4,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        story = []

        def jp_style(size=10, bold=False, align="LEFT", color=colors.black):
            return ParagraphStyle(
                "jp",
                fontName=JP_FONT,
                fontSize=size,
                leading=size * 1.4,
                alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align, 0),
                textColor=color,
                spaceAfter=2,
            )

        # ── タイトル ──
        story.append(Paragraph("御　見　積　書", jp_style(22, bold=True, align="CENTER")))
        story.append(Spacer(1, 8 * mm))

        # ── 案件情報テーブル ──
        today = datetime.now().strftime("%Y年%m月%d日")
        info_data = [
            ["発行日",    today,        "担当者", sales_rep or ""],
            ["お客様名",  f"{client_name} 御中" if client_name else "", "", ""],
            ["現場住所",  site_address or "", "", ""],
        ]
        info_table = Table(info_data, colWidths=[28 * mm, 75 * mm, 22 * mm, 55 * mm])
        info_table.setStyle(TableStyle([
            ("FONT",      (0, 0), (-1, -1), JP_FONT, 9),
            ("BACKGROUND",(0, 0), (0, -1), colors.HexColor("#E3F2FD")),
            ("BACKGROUND",(2, 0), (2, 0),  colors.HexColor("#E3F2FD")),
            ("GRID",      (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",     (0, 0), (0, -1),  "CENTER"),
            ("ALIGN",     (2, 0), (2, 0),   "CENTER"),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 6 * mm))

        # ── 合計金額 ──
        total = estimation.get("total", 0)
        total_data = [[f"御見積合計金額（税込）　¥{total:,}"]]
        total_table = Table(total_data, colWidths=[180 * mm])
        total_table.setStyle(TableStyle([
            ("FONT",       (0, 0), (-1, -1), JP_FONT, 16),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E8F5E9")),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("BOX",        (0, 0), (-1, -1), 1.5, colors.HexColor("#4CAF50")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(total_table)
        story.append(Spacer(1, 6 * mm))

        # ── 明細テーブル ──
        header = ["No", "工種", "品目", "数量", "単位", "単価", "金額", "状態"]
        col_w = [10*mm, 25*mm, 40*mm, 12*mm, 10*mm, 20*mm, 22*mm, 15*mm]

        table_data = [header]
        items = estimation.get("estimation_items", [])
        for i, item in enumerate(items, 1):
            estimated = item.get("estimated", False)
            needs_conf = item.get("needs_confirmation", False)
            status = "⚠️要確認" if needs_conf else ("📊推定" if estimated else "✅確定")
            table_data.append([
                str(i),
                item.get("category", ""),
                item.get("item_name", ""),
                str(item.get("quantity", 0)),
                item.get("unit", ""),
                f"¥{item.get('unit_price', 0):,}",
                f"¥{item.get('amount', 0):,}",
                status,
            ])

        # 小計・税・合計行
        table_data.append(["", "", "", "", "", "小計（税抜）", f"¥{estimation.get('subtotal', 0):,}", ""])
        table_data.append(["", "", "", "", "", "消費税（10%）", f"¥{estimation.get('tax_amount', 0):,}", ""])
        table_data.append(["", "", "", "", "", "合計（税込）", f"¥{total:,}", ""])

        detail_table = Table(table_data, colWidths=col_w, repeatRows=1)

        ts = TableStyle([
            ("FONT",        (0, 0), (-1, -1), JP_FONT, 8),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1A237E")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",       (2, 1), (2, -1),  "LEFT"),
            ("ALIGN",       (5, 1), (6, -1),  "RIGHT"),
            ("GRID",        (0, 0), (-1, -4), 0.5, colors.grey),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, len(items)),
             [colors.white, colors.HexColor("#F5F5F5")]),
            # 小計・税・合計行
            ("BACKGROUND",  (0, -3), (-1, -3), colors.HexColor("#F5F5F5")),
            ("BACKGROUND",  (0, -2), (-1, -2), colors.HexColor("#F5F5F5")),
            ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
            ("FONTSIZE",    (0, -1), (-1, -1), 10),
            ("BOX",         (0, -3), (-1, -1), 0.5, colors.grey),
        ])

        # 推定・要確認行に色付け
        for i, item in enumerate(items, 1):
            if item.get("needs_confirmation"):
                ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFF3E0"))
            elif item.get("estimated"):
                ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E3F2FD"))

        detail_table.setStyle(ts)
        story.append(detail_table)
        story.append(Spacer(1, 4 * mm))

        # ── 注記 ──
        note_style = jp_style(8, color=colors.HexColor("#E65100"))
        story.append(Paragraph(
            "※ 📊推定：現場写真・説明から算出した推定値です。現地確認後に変更になる場合があります。",
            note_style
        ))
        story.append(Paragraph(
            "※ ⚠️要確認：現場での確認が必要な項目です。",
            jp_style(8, color=colors.HexColor("#C62828"))
        ))

        # ── 要確認事項 ──
        confirm_items = estimation.get("confirmation_items", [])
        if confirm_items:
            story.append(Spacer(1, 4 * mm))
            story.append(Paragraph("【要確認事項】", jp_style(9, bold=True)))
            for ci in confirm_items:
                story.append(Paragraph(f"・{ci}", jp_style(9)))

        doc.build(story)
        return str(filename)

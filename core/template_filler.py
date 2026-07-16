"""
テンプレート流し込みモジュール
AIの積算結果をExcelテンプレートの特定セルに書き込む。
計算式（数量×単価=金額 等）はテンプレートのままで保持する。
"""
# redeploy: 2026-06-28 fill_estimation_sheet import 修正 (stale deploy対策)


from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl


# ─────────────────────────────────────────────────────────────
# 標準テンプレート（standard.xlsx）の内訳シート セルマッピング
#
# キー: AIが出力するitem_nameのキーワード（部分一致）
# 値:  (行番号, 数量列D=4, 単価列F=6, 仕様列C=3)
# ─────────────────────────────────────────────────────────────
STANDARD_NAIYAKU_MAPPING = {
    # ── 仮設工事（行3〜10）──
    "外部足場":             (3,  True, True, False),
    "屋根足場":             (4,  True, True, False),
    "昇降設備":             (5,  True, True, False),
    "運搬費":               (6,  True, True, False),
    "道路使用":             (7,  True, True, False),
    "ガードマン":           (8,  True, True, False),
    "カーポート":           (9,  True, True, False),
    "防護管":               (10, True, True, False),
    # ── 塗装工事（行20〜34）──
    "屋根高圧洗浄":         (20, True, True, False),
    "屋根板金":             (21, True, True, True),
    "屋根塗装":             (22, True, True, True),
    "縁切り":               (23, True, True, False),
    "外壁高圧洗浄":         (24, True, True, False),
    "外壁塗装":             (25, True, True, True),
    "土台水切":             (26, True, True, True),
    "中間水切":             (26, True, True, True),  # 土台と同じ行
    "出窓天端":             (27, True, True, True),
    "化粧梁":               (28, True, True, True),
    "付梁":                 (28, True, True, True),
    "軒天塗装（破風m合わせ）": (30, True, True, True),  # soffit_estimate_m のみ行30へ
    # ※ "破風"より前に配置必須: "破風" キーワードが item_name "軒天塗装（破風m合わせ）" に先行マッチするのを防ぐ
    # 行31（玄関庇+ベランダ合算）は fill_standard_template() 内の専用ロジックで書き込む
    # 旧: "軒天（玄関"/"軒天（バルコニー" → 行31 は "軒天塗装" に先行マッチしデッドコードだったため削除
    "破風":                 (29, True, True, True),
    "鼻隠":                 (29, True, True, True),
    "雨樋塗装":             (32, True, True, True),
    "シャッターボックス":   (33, True, True, True),
    "基礎塗装":             (34, True, True, False),
    # ── シーリング工事（行37〜39）──
    "目地シーリング":       (37, True, True, True),
    "サイディング目地":     (37, True, True, True),
    "雑シーリング":         (38, True, True, False),
    "開口部廻りシーリング": (38, True, True, False),
    "トップライト":         (39, True, True, False),
    # ── 諸経費（行42）──
    "諸経費":               (42, False, True, False),  # 数量は1固定・単価のみ
}


def _find_row_for_item(item_name: str) -> Optional[tuple]:
    """item_nameのキーワードでマッピング行を探す"""
    for keyword, mapping in STANDARD_NAIYAKU_MAPPING.items():
        if keyword in item_name:
            return mapping
    return None


def fill_standard_template(
    template_path: Path,
    output_path: Path,
    estimation: dict,
    project_data: dict,
    client_name: str = "",
    site_address: str = "",
    sales_rep: str = "",
    company_name: str = "",
    discount: int = 0,
) -> Path:
    """
    標準テンプレートにAI積算結果を流し込む

    Args:
        template_path: テンプレートExcelのパス
        output_path:   出力先パス
        estimation:    EstimationEngineが返したdict
        project_data:  ImageAnalyzerが返した案件情報
        client_name:   お客様名
        site_address:  現場住所
        sales_rep:     担当者名
        company_name:  受注先（発注元の会社名）
        discount:      値引き額（マイナスで入力、例: -7000）

    Returns:
        Path: 書き込み済みファイルのパス
    """
    import shutil
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    # ── 見積書シートへの書き込み ──
    ws_quote = wb["見積書"]

    today = datetime.now()
    ws_quote["H1"] = today                          # 見積日
    ws_quote["H1"].number_format = "yyyy年m月d日"

    if client_name:
        ws_quote["A4"] = client_name                # 提出先（お客様名）
    if site_address:
        ws_quote["H4"] = site_address               # 現場住所
    if client_name:
        ws_quote["H5"] = f"{client_name}邸 外壁塗装工事"  # 工事件名
    if sales_rep:
        ws_quote["H8"] = sales_rep                  # 担当者

    # 値引き（負の数で入力）
    if discount:
        ws_quote["G18"] = discount if discount <= 0 else -abs(discount)

    # 見積書シート B5 サンプル値クリア（テンプレートに「吉田」が残存するため）
    ws_quote["B5"] = None

    # ── 内訳シートへの書き込み ──
    ws_naiyaku = wb["内訳"]

    # テンプレートのD列サンプル値を事前クリア
    # STANDARD_NAIYAKU_MAPPING の has_qty=True 行をすべてクリアすることで
    # 項目追加時の対象漏れ・将来のサンプル値残存を防ぐ
    # D31（軒天 玄関・バルコニー合算行）は MAPPING 外のため明示追加
    _qty_rows_to_clear = (
        {row for _, (row, has_qty, _, _) in STANDARD_NAIYAKU_MAPPING.items() if has_qty}
        | {31}  # 専用ロジック書き込み行（MAPPINGには存在しない）
    )
    for _r in _qty_rows_to_clear:
        ws_naiyaku.cell(row=_r, column=4).value = None

    # AI積算items → 内訳セルに書き込み
    items = estimation.get("estimation_items", [])

    # すでに書き込んだ行を記録（重複防止）
    written_rows = set()

    for item in items:
        item_name = item.get("item_name", "")
        category = item.get("category", "")
        search_name = item_name + category  # 両方使って検索

        mapping = _find_row_for_item(item_name) or _find_row_for_item(category)
        if mapping is None:
            continue

        row_num, has_qty, has_price, has_spec = mapping

        # 同じ行に複数itemが当たった場合は最初の1件のみ
        if row_num in written_rows:
            continue
        written_rows.add(row_num)

        qty = item.get("quantity", 0) or 0
        unit_price = item.get("unit_price", 0) or 0
        spec = item.get("notes", "") or item.get("basis", "") or ""

        if has_qty and qty:
            ws_naiyaku.cell(row=row_num, column=4).value = qty     # D列: 数量
        if has_price and unit_price:
            ws_naiyaku.cell(row=row_num, column=6).value = unit_price  # F列: 単価
        if has_spec and spec:
            # 仕様欄（C列）は既存の仕様を上書きしない（テンプレートの値を優先）
            existing_spec = ws_naiyaku.cell(row=row_num, column=3).value
            if not existing_spec:
                ws_naiyaku.cell(row=row_num, column=3).value = spec

    # 行31（軒天塗装 玄関・バルコニー㎡）: soffit_entrance_sqm + soffit_balcony_sqm を合算して書き込む
    # MAPPINGで行30に先行マッチするため、専用ロジックで書き込む
    entrance_qty = next(
        (i["quantity"] for i in items if "軒天塗装（玄関庇）" in i.get("item_name", "")), 0.0
    )
    balcony_qty = next(
        (i["quantity"] for i in items if "軒天塗装（ベランダ）" in i.get("item_name", "")), 0.0
    )
    combined_31 = round((entrance_qty or 0.0) + (balcony_qty or 0.0), 2)
    if combined_31 > 0:
        ws_naiyaku.cell(row=31, column=4).value = combined_31  # D31: ㎡合算値
    # combined_31 == 0 の場合は None のまま（事前クリア済みで空欄）

    wb.save(output_path)
    return output_path


def fill_estimation_sheet(
    template_path: Path,
    output_path: Path,
    estimation: dict,
    client_name: str = "",
    site_address: str = "",
    sales_rep: str = "",
    company_name: str = "",
    building_type: str = "",
    estimation_sheet_data: dict = None,
) -> Path:
    """
    積算集計表Excelを生成する。

    estimation_sheet_data が渡された場合（4面入力済み）は per-face 詳細を書き込む。
    なければ estimation の summary から B列のみ書き込む（旧動作互換）。
    """
    import shutil
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["積算集計表"]

    # ── ヘッダー ──────────────────────────────────────────
    if estimation_sheet_data:
        hdr = estimation_sheet_data.get("header", {})
        ws["A1"] = hdr.get("client_name", client_name)
        ws["D1"] = hdr.get("site_address", site_address)
        ws["B2"] = hdr.get("building_type", building_type)
        ws["B3"] = hdr.get("roof_type", "")
        ws["F2"] = hdr.get("company", company_name)
        ws["F3"] = hdr.get("sales_rep", sales_rep)
    else:
        ws["A1"] = client_name
        ws["D1"] = site_address
        ws["B2"] = building_type
        ws["F2"] = company_name
        ws["F3"] = sales_rep

    # ── 4面詳細データがある場合 ───────────────────────────
    if estimation_sheet_data:
        from core.estimation_sheet_builder import ROW_MAP, FACE_COLS, FACES

        rows_data = estimation_sheet_data.get("rows", [])
        for row_def in rows_data:
            key   = row_def["key"]
            total = row_def["total"]
            excel_row = ROW_MAP.get(key)
            if excel_row is None:
                continue

            # B列: 総計
            if total:
                ws.cell(row=excel_row, column=2).value = total

            # 面別列
            for f in FACES:
                face_v = row_def["faces"].get(f, {})
                gross   = face_v.get("gross", 0)
                opening = face_v.get("opening", 0)
                net     = face_v.get("net", gross)  # 開口なければ gross = net

                q_col, o_col, n_col, u_col = FACE_COLS[f]

                if gross:
                    ws.cell(row=excel_row, column=q_col).value = gross
                if opening:
                    ws.cell(row=excel_row, column=o_col).value = opening
                if net and net != gross:
                    ws.cell(row=excel_row, column=n_col).value = net
                elif gross:
                    ws.cell(row=excel_row, column=n_col).value = gross

    else:
        # ── 旧動作: estimation の items から B列のみ ──────
        # 軒天行（10/12/13）はv1.0で面別データなし→書き込みスキップ
        # （Estimation Data v1.0 not_applicable/manual対象外）
        OLD_ROW_MAP = {
            5:  ["外部足場", "足場"],
            6:  ["屋根塗装", "屋根"],
            9:  ["破風", "鼻隠"],
            17: ["外壁塗装", "外壁"],
            21: ["土台水切"],
            34: ["雨樋"],
            41: ["目地シーリング", "目地"],
        }
        items = estimation.get("estimation_items", [])
        for row_num, keywords in OLD_ROW_MAP.items():
            for item in items:
                name = item.get("item_name", "")
                if any(kw in name for kw in keywords):
                    qty = item.get("quantity", 0) or 0
                    if qty:
                        ws.cell(row=row_num, column=2).value = qty
                    break

    wb.save(output_path)
    return output_path

    return output_path


def fill_template(
    template_id: str,
    template_path: Path,
    output_path: Path,
    estimation: dict,
    project_data: dict,
    client_name: str = "",
    site_address: str = "",
    sales_rep: str = "",
    company_name: str = "",
    discount: int = 0,
) -> Path:
    """
    テンプレートIDに応じた流し込み処理を実行するディスパッチャ
    将来的に複数テンプレートに対応するためのエントリポイント
    """
    # 現状は standard のみ対応。将来的にIDで分岐
    return fill_standard_template(
        template_path=template_path,
        output_path=output_path,
        estimation=estimation,
        project_data=project_data,
        client_name=client_name,
        site_address=site_address,
        sales_rep=sales_rep,
        company_name=company_name,
        discount=discount,
    )

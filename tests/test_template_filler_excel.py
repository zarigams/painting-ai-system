"""
template_filler（見積Excel出力）の回帰テスト

対象バグ（2026-07-29 VPS移行Phase C STEP C3-6 で発見）:
  1. テンプレート standard.xlsx の見積書!G18 に元案件の値引き -7015 が残存しており、
     値引き0円のとき fill_standard_template が G18 に触らないため素通りしていた
  2. 内訳シートG列が生値数式（=D×F）のため、明細ごとに円単位へ丸めるアプリ画面の
     小計と、小数点以下を丸めず合計するExcel再計算値が1円単位でずれていた

修正方針（Teppeiさん承認・2026-07-29）:
  - G18は毎回必ず上書き（値引きあり=マイナス値／値引き0円=空欄）
  - 内訳G列にはアプリ側で丸め済みの明細金額(amount)を整数値として書き込む
"""

import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.template_filler import fill_standard_template  # noqa: E402

TEMPLATE_PATH = Path(__file__).parent.parent / "data" / "templates" / "standard.xlsx"


def _item(name, qty, unit_price, category="塗装工事", unit="㎡"):
    """quantity_calculator._item と同じ丸め規則で明細を作るテスト用ヘルパー"""
    amount = round(float(qty) * float(unit_price))
    return {
        "category": category,
        "item_name": name,
        "quantity": float(qty),
        "unit": unit,
        "unit_price": int(unit_price),
        "amount": amount,
        "estimated": False,
        "needs_confirmation": False,
        "notes": "",
    }


def _fill(estimation, discount=0):
    """一時ディレクトリへ流し込み、開き直したworkbookを返す"""
    tmpdir = tempfile.mkdtemp()
    output = Path(tmpdir) / "out.xlsx"
    fill_standard_template(
        template_path=TEMPLATE_PATH,
        output_path=output,
        estimation=estimation,
        project_data={},
        client_name="テスト",
        site_address="東京都",
        sales_rep="担当",
        discount=discount,
    )
    return openpyxl.load_workbook(output)


def _empty_estimation():
    return {"estimation_items": [], "subtotal": 0, "tax_amount": 0, "total": 0}


# ─────────────────────────────────────────────────────────────
# 前提確認: テンプレート自体にG18=-7015が残存している（このテストの意味の担保）
# ─────────────────────────────────────────────────────────────
def test_template_has_leftover_discount_in_g18():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    assert wb["見積書"]["G18"].value == -7015, (
        "テンプレートのG18残存値が変わっている場合、本テスト群の前提を見直すこと"
    )


# ─────────────────────────────────────────────────────────────
# 回帰1: 値引き0円ならG18は空欄になる（テンプレ残存値-7015を素通りさせない）
# ─────────────────────────────────────────────────────────────
def test_g18_cleared_when_no_discount():
    wb = _fill(_empty_estimation(), discount=0)
    assert wb["見積書"]["G18"].value is None


# ─────────────────────────────────────────────────────────────
# 回帰2: 値引き指定時は正しいマイナス値になる
# ─────────────────────────────────────────────────────────────
def test_g18_negative_when_discount_given():
    wb = _fill(_empty_estimation(), discount=50000)
    assert wb["見積書"]["G18"].value == -50000


def test_g18_negative_input_kept_negative():
    wb = _fill(_empty_estimation(), discount=-30000)
    assert wb["見積書"]["G18"].value == -30000


# ─────────────────────────────────────────────────────────────
# 回帰3: 実測ケース相当 — 明細ごと丸めの小計1,979,451円・税込2,177,396円が
#        Excelに書き込まれる金額群と完全一致する
#        （生値合計は1,979,450円台＝丸めタイミング差が実在する構成にしてある）
# ─────────────────────────────────────────────────────────────
def test_totals_match_app_rounding():
    items = [
        _item("外壁塗装", 237.85, 4201),   # 生値  999,207.85 → 丸め   999,208
        _item("屋根塗装", 189.93, 2201),   # 生値  418,035.93 → 丸め   418,036
        _item("外部足場", 350.55, 751),    # 生値  263,263.05 → 丸め   263,263
        _item("諸経費",   1,      298944, category="諸経費", unit="式"),
    ]
    raw_sum = sum(i["quantity"] * i["unit_price"] for i in items)
    app_subtotal = sum(i["amount"] for i in items)

    # 前提の担保: 生値合計と丸め済み合計が実際に食い違う構成であること
    assert app_subtotal == 1_979_451
    assert int(raw_sum) == 1_979_450  # 生値は 1,979,450.83

    tax = round(app_subtotal * 0.10)
    total = app_subtotal + tax
    assert total == 2_177_396

    estimation = {
        "estimation_items": items,
        "subtotal": app_subtotal,
        "tax_amount": tax,
        "total": total,
        "discount": 0,
    }
    wb = _fill(estimation, discount=0)
    ws = wb["内訳"]

    # 書き込まれた明細行のG列（金額）を収集
    written_g = {}
    for row in (25, 22, 3, 42):  # 外壁塗装/屋根塗装/外部足場/諸経費
        written_g[row] = ws.cell(row=row, column=7).value

    # 各G値がアプリ丸め済みamountと一致（数式ではなく整数値であること）
    assert written_g[25] == 999_208
    assert written_g[22] == 418_036
    assert written_g[3] == 263_263
    assert written_g[42] == 298_944
    for v in written_g.values():
        assert isinstance(v, int)

    # Excelが合計する対象（書き込んだG値の和）＝アプリ小計、が完全一致
    assert sum(written_g.values()) == app_subtotal
    # G18は空欄なので、Excel再計算後の税込もアプリと一致する
    # （SUBTOTAL/SUM数式は整数の加算のみで丸め誤差を生まない）
    assert wb["見積書"]["G18"].value is None
    assert round(sum(written_g.values()) * 0.10) + sum(written_g.values()) == total


# ─────────────────────────────────────────────────────────────
# 回帰4: 0.5円が発生する明細でもアプリのamountとExcelのG値が一致する
#        （Pythonの偶数丸めの結果がそのままExcelにも書かれる＝両者は常に同値）
# ─────────────────────────────────────────────────────────────
def test_half_yen_item_matches_app_amount():
    # 0.5×997 = 498.5円 → Pythonのround()は498（偶数丸め）
    item = _item("雑シーリング", 0.5, 997, category="シーリング工事", unit="式")
    assert item["amount"] == round(498.5)  # アプリ側の丸め結果そのもの

    estimation = {
        "estimation_items": [item],
        "subtotal": item["amount"],
        "tax_amount": round(item["amount"] * 0.10),
        "total": item["amount"] + round(item["amount"] * 0.10),
        "discount": 0,
    }
    wb = _fill(estimation, discount=0)
    g = wb["内訳"].cell(row=38, column=7).value  # 雑シーリング → 行38
    assert g == item["amount"]
    assert isinstance(g, int)


# ─────────────────────────────────────────────────────────────
# 回帰5: 既存テンプレートの書式・他セルが壊れない
# ─────────────────────────────────────────────────────────────
def test_template_structure_preserved():
    items = [_item("外壁塗装", 100.0, 4200)]
    estimation = {
        "estimation_items": items,
        "subtotal": items[0]["amount"],
        "tax_amount": round(items[0]["amount"] * 0.10),
        "total": items[0]["amount"] + round(items[0]["amount"] * 0.10),
        "discount": 0,
    }
    wb = _fill(estimation, discount=0)

    # シート構成が維持されている
    assert wb.sheetnames == ["工事概要", "見積書", "内訳", "請求書", "指示書", "予算管理"]

    ws_quote = wb["見積書"]
    # 合計まわりの数式は一切変更しない（テンプレートのまま）
    assert ws_quote["G11"].value == "=内訳!G11+内訳!G35+内訳!G40+内訳!G42"
    assert ws_quote["G19"].value == "=SUBTOTAL(9,G$11:G18)"
    assert ws_quote["G20"].value == "=G19*D20/100"
    assert ws_quote["G21"].value == "=SUBTOTAL(9,G$11:G20)"
    # 値引きのラベル・税率セルは無傷
    assert ws_quote["B18"].value == "値引き"
    assert ws_quote["D20"].value == 10
    # 見積日の表示書式が設定されている
    assert ws_quote["H1"].number_format == "yyyy年m月d日"

    ws_n = wb["内訳"]
    # 書き込んでいない行のG列数式はテンプレートのまま（=D×F）
    assert ws_n.cell(row=12, column=7).value == "=D12*F12"
    assert ws_n.cell(row=20, column=7).value == "=D20*F20"  # 屋根高圧洗浄（未書込）
    # カテゴリ小計・カテゴリ見出しの数式/ラベルも無傷
    assert ws_n.cell(row=11, column=7).value == "=SUM(G3:G10)"
    assert ws_n.cell(row=35, column=7).value == "=SUM(G20:G34)"
    assert ws_n.cell(row=37, column=2).value == "サイデイング目地シーリング"
    # 書き込んだ行（外壁塗装=行25）はD/F/Gが更新されている
    assert ws_n.cell(row=25, column=4).value == 100.0
    assert ws_n.cell(row=25, column=6).value == 4200
    assert ws_n.cell(row=25, column=7).value == items[0]["amount"]


# ─────────────────────────────────────────────────────────────
# 回帰5補: 行31（軒天 玄関庇+ベランダ合算）のG値は
#          「丸め済みamountの和」（合算㎡×単価の再丸めではない）
# ─────────────────────────────────────────────────────────────
def test_row31_amount_is_sum_of_rounded_amounts():
    e = _item("軒天塗装（玄関庇）", 3.55, 951)   # 3,376.05 → 3,376
    b = _item("軒天塗装（ベランダ）", 4.15, 951)  # 3,946.65 → 3,947
    estimation = {
        "estimation_items": [e, b],
        "subtotal": e["amount"] + b["amount"],
        "tax_amount": 0,
        "total": e["amount"] + b["amount"],
        "discount": 0,
    }
    wb = _fill(estimation, discount=0)
    ws_n = wb["内訳"]
    assert ws_n.cell(row=31, column=4).value == round(3.55 + 4.15, 2)
    # 個別丸めの和 3,376+3,947=7,323（合算再計算だと round(7.7×951)=7,323 だが、
    # 定義として「amountの和」であることを確認する）
    assert ws_n.cell(row=31, column=7).value == e["amount"] + b["amount"] == 7_323
